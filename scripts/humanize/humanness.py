"""Humanness scoring adapter for BioPhi (Merck) / Sapiens.

BioPhi (github.com/Merck/BioPhi) provides:
  * Sapiens: an antibody language model producing per-position probability
    matrices; `--mean-score-only` yields one humanness score per sequence.
  * OASis: humanness identity vs human antibody repertoires (needs the
    large OASis 9-mer DB, server-side).

This adapter runs biophi on our variant sequences and reports the mean
Sapiens score per variant (and OASis identity when the DB is available) as
an independent humanness cross-check of our germline-identity metric.

Setup (server):
    conda create -n biophi python=3.9
    conda install -n biophi biophi -c bioconda -c conda-forge --override-channels
    # OASis DB (optional, 22 GB): wget https://zenodo.org/record/5164685/files/OASis_9mers_v1.db.gz

Usage:
    humanize run --input seq.fasta --biophi-env biophi \
                 [--oasis-db /path/OASis_9mers_v1.db]
"""

from __future__ import annotations

import csv
import os
import shutil
import subprocess
import tempfile
from dataclasses import dataclass, field
from typing import Dict, List, Optional


@dataclass
class HumannessResult:
    sapiens_mean: Optional[float] = None
    oasis_identity: Optional[float] = None
    n_sequences: int = 0
    source: str = "biophi"
    note: str = ""


def _run(cmd: List[str], timeout: int = 600) -> subprocess.CompletedProcess:
    return subprocess.run(cmd, capture_output=True, text=True, timeout=timeout)


def _conda_run(env: str, cmd: List[str], timeout: int = 600) -> subprocess.CompletedProcess:
    """Run a command inside a conda env (best-effort; falls back to PATH)."""
    if shutil.which(cmd[0]):
        return _run(cmd, timeout)
    base = os.environ.get("CONDA_EXE") or shutil.which("conda")
    if not base:
        raise RuntimeError("biophi not on PATH and conda not found")
    full = [base, "run", "-n", env, *cmd]
    return _run(full, timeout)


def run_sapiens(sequences: Dict[str, str], env: str = "biophi",
                binary: str = "biophi") -> Dict[str, HumannessResult]:
    """Mean Sapiens humanness score per sequence.

    sequences: {name: aa_sequence}
    Returns {name: HumannessResult(sapiens_mean=...)}.
    """
    out = {}
    with tempfile.TemporaryDirectory() as td:
        fasta = os.path.join(td, "input.fa")
        with open(fasta, "w") as fh:
            for name, seq in sequences.items():
                fh.write(f">{name}\n{seq}\n")
        try:
            proc = _conda_run(env, [binary, "sapiens", fasta, "--mean-score-only"])
        except (RuntimeError, FileNotFoundError) as e:
            return {n: HumannessResult(note=f"biophi unavailable: {e}") for n in sequences}
        if proc.returncode != 0:
            return {n: HumannessResult(note=f"biophi failed: {proc.stderr[-200:]}") for n in sequences}
        # output CSV: sequence,mean_score (or a file in outdir)
        out_csv = os.path.join(td, "scores.csv")
        if os.path.exists(out_csv):
            csv_path = out_csv
        else:
            csv_path = td  # biophi may write to --output dir; scan for csv
            candidates = [os.path.join(td, f) for f in os.listdir(td) if f.endswith(".csv")]
            csv_path = candidates[0] if candidates else None
        if csv_path:
            with open(csv_path) as fh:
                for row in csv.DictReader(fh):
                    name = row.get("sequence", "") or row.get("id", "")
                    score = row.get("mean_score") or row.get("score")
                    if name in sequences and score is not None:
                        try:
                            out[name] = HumannessResult(
                                sapiens_mean=round(float(score), 4), n_sequences=1)
                        except ValueError:
                            pass
    return out


def run_oasis(sequences: Dict[str, str], oasis_db: str,
              env: str = "biophi") -> Dict[str, HumannessResult]:
    """OASis identity per sequence (requires the 22 GB OASis 9-mer DB)."""
    out = {}
    if not os.path.exists(oasis_db):
        return {n: HumannessResult(note="OASis DB not found") for n in sequences}
    with tempfile.TemporaryDirectory() as td:
        fasta = os.path.join(td, "input.fa")
        with open(fasta, "w") as fh:
            for name, seq in sequences.items():
                fh.write(f">{name}\n{seq}\n")
        try:
            proc = _conda_run(env, ["biophi", "oasis", fasta,
                                    "--oasis-db", oasis_db,
                                    "--output", os.path.join(td, "oasis.xlsx")])
        except (RuntimeError, FileNotFoundError) as e:
            return {n: HumannessResult(note=f"biophi unavailable: {e}") for n in sequences}
        if proc.returncode != 0:
            return {n: HumannessResult(note=f"oasis failed: {proc.stderr[-200:]}") for n in sequences}
        # xlsx parsing requires openpyxl; keep it optional
        try:
            import openpyxl  # type: ignore
            wb = openpyxl.load_workbook(os.path.join(td, "oasis.xlsx"))  # type: ignore
            ws = wb.active
            if ws is not None:
                header = [c.value for c in next(ws.iter_rows())]
                for row in ws.iter_rows(min_row=2):
                    d = dict(zip(header, [c.value for c in row]))
                    name = d.get("sequence") or d.get("id")
                    ident = d.get("identity")
                    if name in sequences and ident is not None:
                        try:
                            out[name] = HumannessResult(oasis_identity=round(float(ident), 4))
                        except (TypeError, ValueError):
                            pass
        except ImportError:
            out = {n: HumannessResult(note="oasis output requires openpyxl") for n in sequences}
    return out
